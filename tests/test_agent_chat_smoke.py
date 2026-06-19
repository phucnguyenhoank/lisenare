"""
Smoke test for POST /agent/chat.

Pytest mode (default):
    pytest tests/test_agent_chat_smoke.py -v

Script mode (writes results back to xlsx):
    python tests/test_agent_chat_smoke.py
    python tests/test_agent_chat_smoke.py --output tests/smoke_results.xlsx
"""

import sys
import argparse
from datetime import datetime
from pathlib import Path

# Ensure project root is on sys.path when running as a script
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pytest
from sqlalchemy.exc import SQLAlchemyError
from sqlmodel import Session, text
from fastapi.testclient import TestClient

from app.main import app
from app.database import engine

XLSX_PATH = Path(__file__).parent / "chat_test_questions.xlsx"
SHEET_NAME = "agent_chat"
ANSWER_PREVIEW_LEN = 120

# ── Styles ────────────────────────────────────────────────────────────────────
_PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
_SKIP_FILL = PatternFill("solid", fgColor="FFEB9C")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


# ── Data loading ──────────────────────────────────────────────────────────────

def _load_cases():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num, _, learner_id, message, _, tool_expected, multi_tool_str, _ = row
        if row_num is None:
            continue
        cases.append({
            "row_num": int(row_num),
            "learner_id": int(learner_id),
            "message": message,
            "tool_expected": tool_expected,
            "multi_tool": multi_tool_str == "Yes",
        })
    wb.close()
    return cases


def load_agent_chat_cases():
    return [
        pytest.param(
            c["learner_id"], c["message"], c["tool_expected"], c["multi_tool"],
            id=f"#{c['row_num']} {c['tool_expected']}",
        )
        for c in _load_cases()
    ]


# ── Core call (shared by pytest and script modes) ─────────────────────────────

def _run_case(client: TestClient, learner_id: int, message: str, tool_expected: str):
    """Returns (passed: bool, tool_calls_str: str, answer_preview: str, error: str)."""
    try:
        response = client.post(
            "/agent/chat",
            json={
                "learner_id": learner_id,
                "messages": [{"role": "user", "content": message}],
            },
        )
        if response.status_code != 200:
            return False, "", "", f"HTTP {response.status_code}: {response.text[:200]}"

        data = response.json()
        answer = data.get("answer", "")
        tool_calls = data.get("tool_calls", [])
        called_names = [tc["name"] for tc in tool_calls]
        tools_str = ", ".join(called_names) if called_names else "(none)"
        preview = answer.strip()[:ANSWER_PREVIEW_LEN]
        full_answer = answer.strip()

        if not answer.strip():
            return False, tools_str, preview, "answer is empty", full_answer
        if tool_expected != "no_tool" and tool_expected not in called_names:
            return False, tools_str, preview, f"expected '{tool_expected}' not in {called_names}", full_answer

        return True, tools_str, preview, "", full_answer
    except Exception as exc:
        return False, "", "", str(exc), ""


# ── Pytest mode ───────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def agent_client():
    try:
        with Session(engine) as s:
            s.exec(text("SELECT 1"))
    except SQLAlchemyError as exc:
        pytest.skip(f"DB unavailable: {exc}")
    with TestClient(app) as client:
        yield client


@pytest.mark.parametrize(
    "learner_id,message,tool_expected,multi_tool",
    load_agent_chat_cases(),
)
def test_agent_chat_smoke(agent_client, learner_id, message, tool_expected, multi_tool):
    passed, tools_str, preview, error = _run_case(
        agent_client, learner_id, message, tool_expected
    )
    assert passed, error


# ── Script mode ───────────────────────────────────────────────────────────────

def _init_result_columns(ws):
    """Write result column headers (cols I–L) and set widths. Idempotent."""
    result_headers = ["Tools called", "Pass/Fail", "Answer preview", "Error", "Full answer"]
    for i, h in enumerate(result_headers, start=9):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.column_dimensions["I"].width = 35
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 55
    ws.column_dimensions["L"].width = 45
    ws.column_dimensions["M"].width = 80


def _write_row(ws, row_num: int, passed: bool, tools_str: str, preview: str, error: str, full_answer: str = ""):
    xlsx_row = row_num + 1  # +1 for header
    fill = _PASS_FILL if passed else _FAIL_FILL
    ws.cell(row=xlsx_row, column=9, value=tools_str).fill = fill
    pf_cell = ws.cell(row=xlsx_row, column=10, value="PASS" if passed else "FAIL")
    pf_cell.fill = fill
    pf_cell.font = Font(bold=True, color="375623" if passed else "9C0006")
    ws.cell(row=xlsx_row, column=11, value=preview).fill = fill
    ws.cell(row=xlsx_row, column=12, value=error).fill = fill
    ws.cell(row=xlsx_row, column=13, value=full_answer).fill = fill
    for col in range(9, 14):
        ws.cell(row=xlsx_row, column=col).alignment = Alignment(
            wrap_text=True, vertical="top"
        )


def _load_done_rows(out_path: Path) -> set[int]:
    """Return set of row_num already tested (col J has PASS or FAIL)."""
    if not out_path.exists():
        return set()
    wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num = row[0]
        pass_fail = row[9] if len(row) > 9 else None  # col J (0-indexed = 9)
        if row_num is not None and pass_fail in ("PASS", "FAIL"):
            done.add(int(row_num))
    wb.close()
    return done


def _run_script(out_path: Path):
    try:
        with Session(engine) as s:
            s.exec(text("SELECT 1"))
    except SQLAlchemyError as exc:
        print(f"DB unavailable, aborting: {exc}")
        sys.exit(1)

    import shutil
    if not out_path.exists():
        shutil.copy(XLSX_PATH, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb[SHEET_NAME]
    _init_result_columns(ws)
    wb.save(out_path)

    done_rows = _load_done_rows(out_path)
    all_cases = _load_cases()
    pending = [c for c in all_cases if c["row_num"] not in done_rows]

    if done_rows:
        print(f"Resuming: {len(done_rows)} already done, {len(pending)} remaining.")
    passed_count = sum(1 for c in all_cases if c["row_num"] in done_rows)

    with TestClient(app) as client:
        for c in pending:
            passed, tools_str, preview, error, full_answer = _run_case(
                client, c["learner_id"], c["message"], c["tool_expected"]
            )
            if passed:
                passed_count += 1
            status = "PASS" if passed else "FAIL"
            print(f"[{status}] #{c['row_num']:>3} {c['tool_expected']:<32} | {tools_str}")
            if error:
                print(f"       ERROR: {error}")

            # Write and save immediately so results survive any crash or interrupt
            wb = openpyxl.load_workbook(out_path)
            ws = wb[SHEET_NAME]
            _write_row(ws, c["row_num"], passed, tools_str, preview, error, full_answer)
            wb.save(out_path)

    total = len(all_cases)
    print(f"\nResults: {passed_count}/{total} passed")
    print(f"Saved  : {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output", type=Path,
        default=Path(__file__).parent / f"smoke_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    args = parser.parse_args()
    _run_script(args.output)
