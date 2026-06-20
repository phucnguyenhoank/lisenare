"""
Regenerate ALL 100 rows in the grammar_chat sheet with full question lists
fetched from the database for each exercise.

Run: python tests/generate_grammar_questions.py
"""

import json
import random
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlmodel import Session, select

from app.database import engine
from app.database.models import Exercise, Question

XLSX_PATH = Path(__file__).parent / "chat_test_questions.xlsx"
SHEET_NAME = "grammar_chat"

EXERCISE_IDS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]

# 10 user messages per exercise = 100 total cases
QUESTIONS_BY_EXERCISE = {
    1: [
        (4, "Cho tôi biết quy tắc thêm -s/-es vào động từ thì hiện tại đơn."),
        (2, "Khi nào dùng 'do' và 'does' trong câu hỏi hiện tại đơn?"),
        (3, "Câu phủ định thì hiện tại đơn được tạo ra như thế nào?"),
        (4, "Giải thích sự khác biệt giữa 'I eat' và 'I am eating'."),
        (2, "Tôi điền sai rồi, đáp án đúng là gì và tại sao?"),
        (3, "Câu 3 tôi chưa hiểu, giải thích giúp tôi."),
        (4, "Cho tôi thêm ví dụ về thì hiện tại đơn với chủ ngữ số nhiều."),
        (2, "Tại sao câu 2 đáp án là 'goes' không phải 'go'?"),
        (3, "Adverbs of frequency đặt ở đâu trong câu hiện tại đơn?"),
        (4, "Tôi còn 2 câu chưa trả lời, bắt đầu từ đâu?"),
    ],
    2: [
        (3, "Quy tắc chia động từ ở thì quá khứ đơn là gì?"),
        (
            4,
            "Động từ bất quy tắc thì quá khứ đơn — cho tôi danh sách phổ biến.",
        ),
        (2, "Câu hỏi thì quá khứ đơn dùng 'did' như thế nào?"),
        (3, "Sự khác biệt giữa past simple và past continuous là gì?"),
        (4, "Tôi đặt 'yesterday' ở đâu trong câu?"),
        (2, "Câu 4 tôi trả lời sai, giải thích lỗi sai của tôi."),
        (3, "Khi nào dùng past simple, khi nào dùng present perfect?"),
        (4, "Động từ 'go' ở quá khứ là gì?"),
        (2, "Câu 1 và câu 3 liên quan đến nhau như thế nào?"),
        (3, "Tôi không hiểu câu 5, hướng dẫn từng bước."),
    ],
    3: [
        (2, "Present continuous dùng khi nào, cho ví dụ thực tế."),
        (3, "Tại sao một số động từ không dùng được ở dạng continuous?"),
        (4, "Cách tạo câu phủ định ở thì hiện tại tiếp diễn."),
        (2, "Phân biệt 'I think' và 'I am thinking'."),
        (3, "Cách đặt câu hỏi với 'What are you doing?'"),
        (4, "Câu 2 tôi điền 'plays' nhưng sai, lý do tại sao?"),
        (2, "Signal words của present continuous là gì?"),
        (3, "Tại sao không nói 'I am knowing the answer'?"),
        (4, "Giải thích cách dùng '-ing' với động từ to be."),
        (2, "Câu 5 khó quá, gợi ý cho tôi."),
    ],
    4: [
        (4, "Will và be going to khác nhau như thế nào?"),
        (2, "Khi nào dùng present continuous để nói về tương lai?"),
        (3, "Câu điều kiện loại 1 liên quan đến future simple như thế nào?"),
        (4, "Giải thích 'shall' trong câu tương lai."),
        (2, "Tôi muốn nói 'tôi sẽ gặp bạn vào ngày mai' bằng tiếng Anh."),
        (3, "Câu 3 tôi chưa chắc, kiểm tra đáp án giúp tôi."),
        (4, "Future simple dùng với dự đoán như thế nào?"),
        (2, "Tại sao câu 1 dùng 'will' mà không dùng 'going to'?"),
        (3, "Cho ví dụ câu tương lai với 'probably'."),
        (4, "Câu 4 và câu 5 khác nhau điểm gì?"),
    ],
    5: [
        (3, "Sự khác biệt giữa 'must' và 'have to'."),
        (4, "Khi nào dùng 'should' thay vì 'must'?"),
        (2, "Modal verbs không chia theo ngôi — giải thích tại sao."),
        (3, "Cách tạo câu phủ định với modal verbs."),
        (4, "Phân biệt 'can' và 'could' trong xin phép."),
        (2, "Câu 2 tôi trả lời 'musts' nhưng sai, tại sao?"),
        (3, "Modal perfect (must have, should have) nghĩa là gì?"),
        (4, "Câu 4 dùng modal nào là đúng nhất?"),
        (2, "Cách dùng 'would' để diễn đạt thói quen trong quá khứ."),
        (3, "Tôi còn 3 câu chưa làm, câu nào dễ nhất?"),
    ],
    6: [
        (2, "Giải thích cách dùng 'in', 'on', 'at' với thời gian."),
        (
            3,
            "Preposition sau động từ — tại sao 'interested in' chứ không phải 'interested at'?",
        ),
        (4, "Sự khác biệt giữa 'in the morning' và 'at morning'."),
        (2, "Cho tôi danh sách các cụm từ preposition hay dùng."),
        (3, "Tại sao câu này sai: 'She arrived to school late'?"),
        (4, "Câu 3 tôi điền 'at' nhưng sai, preposition đúng là gì?"),
        (2, "'On time' và 'in time' khác nhau như thế nào?"),
        (3, "Preposition of place: 'in', 'on', 'at' với địa điểm."),
        (4, "Câu 5 liên quan đến preposition nào?"),
        (2, "Tại sao nói 'depend on' không phải 'depend of'?"),
    ],
    7: [
        (4, "Khi nào dùng 'a' và khi nào dùng 'an'?"),
        (2, "Giải thích khi nào không cần dùng article."),
        (3, "Sự khác biệt giữa 'a dog' và 'the dog'."),
        (4, "Tại sao nói 'the sun' nhưng không nói 'a sun'?"),
        (2, "Articles với tên riêng và địa danh — quy tắc như thế nào?"),
        (3, "Câu 2 tôi điền 'a' nhưng sai, tại sao phải dùng 'an'?"),
        (4, "Khi nào dùng 'the' với danh từ số nhiều?"),
        (2, "Tại sao nói 'go to school' không có article?"),
        (3, "Câu 4 có cụm từ đặc biệt nào không?"),
        (4, "'A' vs 'one' — khác nhau như thế nào?"),
    ],
    8: [
        (3, "Giải thích 3 loại câu điều kiện chính trong tiếng Anh."),
        (4, "Conditional type 2 khác type 1 như thế nào?"),
        (2, "Câu điều kiện loại 3 nói về điều gì?"),
        (3, "Mixed conditionals là gì và khi nào dùng?"),
        (4, "Tại sao câu này sai: 'If it will rain, I stay home'?"),
        (2, "Câu 3 tôi dùng 'would' sai chỗ, giải thích giúp."),
        (3, "Unless trong câu điều kiện có nghĩa gì?"),
        (4, "Câu 1 dùng conditional type mấy?"),
        (2, "Cấu trúc câu điều kiện loại 2 là gì?"),
        (3, "Tôi không phân biệt được câu 4 và câu 5."),
    ],
    9: [
        (2, "Quy tắc backshift tenses trong reported speech."),
        (3, "Reporting questions — cách đổi câu hỏi sang reported speech."),
        (4, "Khi nào không cần backshift trong reported speech?"),
        (2, "Đổi đại từ nhân xưng trong reported speech như thế nào?"),
        (3, "Say vs tell — sự khác biệt trong reported speech."),
        (4, "Câu 2 tôi dùng 'said me' — đúng hay sai?"),
        (2, "Reporting verbs ngoài 'say' và 'tell' còn có gì?"),
        (3, "Câu 4 cần đổi thời nào sang thời nào?"),
        (4, "Reported speech với câu mệnh lệnh như thế nào?"),
        (2, "Tôi không hiểu câu 3, cho ví dụ tương tự."),
    ],
    10: [
        (
            2,
            "Passive voice ở thì hiện tại tiếp diễn được hình thành như thế nào?",
        ),
        (4, "Khi nào thì không nên dùng câu bị động?"),
        (3, "Get passive vs be passive — khác nhau ở điểm nào?"),
        (2, "Passive voice với modal verbs — ví dụ cụ thể."),
        (
            4,
            "Đổi câu chủ động sau sang bị động: 'People speak English worldwide.'",
        ),
        (3, "Câu 3 tôi điền 'is written' hay 'was written'?"),
        (2, "Passive voice ở past perfect là gì?"),
        (4, "Tại sao câu 1 dùng 'is being built' không phải 'is built'?"),
        (3, "Câu bị động với 2 tân ngữ như thế nào?"),
        (2, "Tôi còn câu 5 chưa làm, gợi ý hướng đi."),
    ],
}


def _fetch_exercise_questions(
    session: Session, exercise_id: int
) -> tuple[str, list[dict]]:
    """Return (exercise_name, list of QuestionContext dicts) from DB."""
    exercise = session.get(Exercise, exercise_id)
    if exercise is None:
        raise ValueError(f"Exercise id={exercise_id} not found in DB")

    qs = session.exec(
        select(Question)
        .where(Question.exercise_id == exercise_id)
        .order_by(Question.id)
    ).all()

    questions = [
        {
            "order_id": i + 1,
            "question_id": q.id,
            "question": q.question or q.content or "",
            "user_answer": None,
        }
        for i, q in enumerate(qs)
    ]
    return exercise.name, questions


def _make_request_json(
    learner_id: int,
    exercise_id: int,
    exercise_name: str,
    message: str,
    questions: list[dict],
    case_index: int,
) -> str:
    # Simulate realistic state: randomly mark ~half the questions as answered
    answered = [q.copy() for q in questions]
    rng = random.Random(learner_id * 1000 + exercise_id * 10 + case_index)
    choices = ["A", "B", "C", "D"]
    for q in answered:
        if rng.random() < 0.5:
            q["user_answer"] = rng.choice(choices)

    body = {
        "learner_id": learner_id,
        "session_id": None,
        "messages": [{"role": "user", "content": message}],
        "context": {
            "exercise_id": exercise_id,
            "exercise_name": exercise_name,
            "current_question_id": None,
            "questions": answered,
        },
    }
    return json.dumps(body, ensure_ascii=False, indent=2)


def _write_header(ws):
    headers = [
        "#",
        "Endpoint",
        "learner_id",
        "exercise_id",
        "exercise_name",
        "Câu hỏi người dùng",
        "Request JSON mẫu",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    # Clear all data rows, keep header
    max_row = ws.max_row
    if max_row >= 2:
        ws.delete_rows(2, max_row - 1)
        print(f"Cleared {max_row - 1} existing data rows.")

    with Session(engine) as db_session:
        row_num = 1
        xlsx_row = 2  # row 1 = header

        for exercise_id in EXERCISE_IDS:
            try:
                exercise_name, questions = _fetch_exercise_questions(
                    db_session, exercise_id
                )
            except ValueError as e:
                print(f"SKIP: {e}")
                continue

            print(
                f"Exercise {exercise_id} '{exercise_name}': {len(questions)} questions in DB"
            )

            for case_idx, (learner_id, message) in enumerate(
                QUESTIONS_BY_EXERCISE[exercise_id]
            ):
                req_json = _make_request_json(
                    learner_id,
                    exercise_id,
                    exercise_name,
                    message,
                    questions,
                    case_idx,
                )
                ws.cell(row=xlsx_row, column=1, value=row_num)
                ws.cell(row=xlsx_row, column=2, value="POST /grammar/chat")
                ws.cell(row=xlsx_row, column=3, value=learner_id)
                ws.cell(row=xlsx_row, column=4, value=exercise_id)
                ws.cell(row=xlsx_row, column=5, value=exercise_name)
                ws.cell(row=xlsx_row, column=6, value=message)
                ws.cell(row=xlsx_row, column=7, value=req_json)
                row_num += 1
                xlsx_row += 1

    wb.save(XLSX_PATH)
    print(f"Done. Written {row_num - 1} rows total.")


if __name__ == "__main__":
    main()
