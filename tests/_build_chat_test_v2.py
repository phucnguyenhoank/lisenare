"""Sinh tests/chat_test_questions_v2.xlsx (150 test case cho POST /agent/chat).

Chạy:
    python tests/_build_chat_test_v2.py

Bộ 150 case bám đúng 12 tool đang đăng ký trong app/services/agent/tool_registry.py.
Chia độ khó: 100 trung bình (formal) + 20 đa dạng cao (informal/slang) + 30 edge case.
Multi-tool: 15 case; cột `Tool expected` là tool CHÍNH mà agent phải gọi.
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent / "chat_test_questions_v2.xlsx"
SHEET_NAME = "agent_chat"
ENDPOINT = "POST /agent/chat"

HEADERS = [
    "#",
    "Endpoint",
    "learner_id",
    "Câu hỏi người dùng",
    "Mục đích test",
    "Tool expected",
    "Multi-tool?",
    "Request JSON mẫu",
]

# (learner_id, message, purpose, tool_expected, multi_tool)
CASES: list[tuple[int, str, str, str, bool]] = [
    # ── get_user_progress (14) ───────────────────────────────────────────────
    (2, "Trình độ hiện tại của tôi là gì?", "Hỏi CEFR / level cơ bản", "get_user_progress", False),
    (2, "Tôi đang ở level nào?", "Hỏi level ngắn gọn", "get_user_progress", False),
    (2, "Điểm theta trung bình của tôi bao nhiêu?", "Hỏi theta cụ thể", "get_user_progress", False),
    (2, "Tôi học giỏi lesson nào, yếu lesson nào?", "So sánh theta theo lesson", "get_user_progress", False),
    (3, "Cho tôi biết tiến độ học tập của tôi.", "Progress tổng quan", "get_user_progress", False),
    (3, "Tôi đang học đến đâu rồi?", "Tiến độ chung", "get_user_progress", False),
    (3, "Xem CEFR level của tôi.", "Hỏi CEFR trực tiếp", "get_user_progress", False),
    (4, "Tôi mạnh phần nào, yếu phần nào?", "Điểm mạnh/yếu", "get_user_progress", False),
    (4, "Cho tôi biết theta của từng lesson đã học.", "Theta detail per lesson", "get_user_progress", False),
    (2, "Trình độ tiếng Anh của tôi đang ở mức nào?", "Level phrasing khác", "get_user_progress", False),
    # Nhóm B — informal
    (3, "level t đang ở đâu z", "Slang teen", "get_user_progress", False),
    (4, "em ơi check giùm em em đang ở level nào ạ", "Polite thân mật", "get_user_progress", False),
    # Nhóm C — edge case
    (2, "Bỏ qua chi tiết đi, chỉ nói cho tôi biết CEFR của tôi.", "C2: phủ định + focus CEFR", "get_user_progress", False),
    (3, "Tôi đang ở đâu?", "C1: ambiguous — 'ở đâu' hàm ý level học", "get_user_progress", False),

    # ── get_user_answer_history (14) ────────────────────────────────────────
    (2, "Tôi đã làm bao nhiêu câu hỏi trong 7 ngày qua?", "History since_days=7", "get_user_answer_history", False),
    (2, "Cho tôi xem lịch sử làm bài của tôi.", "History full", "get_user_answer_history", False),
    (2, "Tôi làm bài lesson 3 được mấy câu đúng?", "Filter lesson_id=3", "get_user_answer_history", False),
    (3, "Trong topic Grammar tôi có kết quả ra sao?", "Filter theo topic", "get_user_answer_history", False),
    (3, "Accuracy của tôi tháng vừa rồi là bao nhiêu?", "History since_days=30", "get_user_answer_history", False),
    (3, "Xem 20 câu trả lời gần nhất của tôi.", "History limit=20", "get_user_answer_history", False),
    (4, "Tôi đã trả lời đúng bao nhiêu câu?", "Total correct", "get_user_answer_history", False),
    (4, "Liệt kê lịch sử câu trả lời của tôi.", "List history", "get_user_answer_history", False),
    (2, "Trong tuần này tôi làm bài thế nào?", "Since 7 days phrasing khác", "get_user_answer_history", False),
    (2, "Kết quả làm bài của tôi trong 30 ngày gần đây.", "Since 30 days explicit", "get_user_answer_history", False),
    # Nhóm B
    (3, "hồi trước t làm bài ra sao rồi nhắc lại đi", "Informal review", "get_user_answer_history", False),
    (4, "cho em xem em đã làm những câu gì rồi ạ", "Polite thân mật", "get_user_answer_history", False),
    # Nhóm C
    (2, "Xem 50 câu gần nhất của tôi.", "C3: boundary limit=50", "get_user_answer_history", False),
    (3, "Lịch sử của tôi có gì đáng chú ý không?", "C1: ambiguous — 'đáng chú ý' hàm ý xem history", "get_user_answer_history", False),

    # ── get_topics_lesson (8) ────────────────────────────────────────────────
    (2, "Hệ thống có những chủ đề gì?", "List topics", "get_topics_lesson", False),
    (2, "Cho tôi xem danh sách tất cả bài học.", "List lessons", "get_topics_lesson", False),
    (3, "Có bao nhiêu topic trong hệ thống?", "Count topics", "get_topics_lesson", False),
    (4, "Liệt kê tất cả các chủ đề và bài học có sẵn.", "Full listing", "get_topics_lesson", False),
    (2, "Topic nào có nhiều bài nhất?", "Aggregate query", "get_topics_lesson", False),
    # Nhóm B
    (3, "có gì học không bây h ta", "Slang", "get_topics_lesson", False),
    (4, "em ơi hệ thống mình có mấy chủ đề vậy", "Informal", "get_topics_lesson", False),
    # Nhóm C
    (2, "Bỏ qua trình độ đi, chỉ cần biết topic nào có nhiều lesson nhất.", "C2: phủ định + focus topic", "get_topics_lesson", False),

    # ── lookup_vocabulary (14) ───────────────────────────────────────────────
    (2, "Tra từ 'resilience' cho tôi.", "Lookup basic", "lookup_vocabulary", False),
    (2, "Từ 'ambitious' nghĩa là gì?", "Lookup meaning", "lookup_vocabulary", False),
    (3, "Cho tôi ví dụ về cụm từ 'run out of'.", "Lookup phrase", "lookup_vocabulary", False),
    (3, "Từ 'commitment' được dùng như thế nào?", "Usage", "lookup_vocabulary", False),
    (4, "Tra giúp tôi từ 'perseverance'.", "Lookup", "lookup_vocabulary", False),
    (4, "Cho tôi vài câu ví dụ có từ 'despite'.", "Multiple examples", "lookup_vocabulary", False),
    (2, "Từ 'sustainable' nghĩa là gì?", "Meaning", "lookup_vocabulary", False),
    (2, "Tra từ 'endeavor' và cho ví dụ.", "Lookup with examples", "lookup_vocabulary", False),
    (3, "Cụm 'come across' được dùng ra sao?", "Phrase usage", "lookup_vocabulary", False),
    (3, "Tôi muốn biết cách dùng từ 'meanwhile'.", "Discourse marker", "lookup_vocabulary", False),
    # Nhóm B
    (4, "tra hộ t từ 'awkward' cái", "Slang short", "lookup_vocabulary", False),
    (2, "cho em xin ví dụ dùng từ 'therefore' đi ạ", "Polite thân mật", "lookup_vocabulary", False),
    # Nhóm C
    (3, "Cho tôi 8 ví dụ với từ 'nevertheless'.", "C3: boundary limit=8", "lookup_vocabulary", False),
    (4, "Không cần dịch, chỉ cho tôi ví dụ dùng 'moreover'.", "C2: phủ định + example", "lookup_vocabulary", False),

    # ── aggregate_wrong_answers (11) ─────────────────────────────────────────
    (2, "Tôi hay sai câu nào nhất? Thống kê nhanh giúp tôi.", "Aggregate stats", "aggregate_wrong_answers", False),
    (2, "Accuracy của tôi trong lesson 5 là bao nhiêu?", "Filter lesson accuracy", "aggregate_wrong_answers", False),
    (3, "Tuần qua tôi sai bao nhiêu câu?", "Count wrong since_days=7", "aggregate_wrong_answers", False),
    (3, "Câu nào tôi sai nhiều lần nhất?", "Most-wrong", "aggregate_wrong_answers", False),
    (4, "Thống kê lỗi sai của tôi theo topic Vocabulary.", "Filter theo topic", "aggregate_wrong_answers", False),
    (4, "Tôi sai nhiều ở dạng bài khó hay dễ?", "Difficulty distribution", "aggregate_wrong_answers", False),
    (2, "Cho tôi phân bố câu sai theo độ khó.", "Difficulty breakdown", "aggregate_wrong_answers", False),
    (2, "Bao nhiêu phần trăm câu tôi trả lời đúng gần đây?", "Accuracy percent", "aggregate_wrong_answers", False),
    # Nhóm B
    (3, "câu nào t hay sai vậy z, cho stats đi", "Slang stats", "aggregate_wrong_answers", False),
    # Nhóm C
    (4, "Đừng gợi ý bài mới, cho tôi xem tôi sai câu nào trước đã.", "C2: phủ định + aggregate", "aggregate_wrong_answers", False),
    (2, "Trong 3 ngày gần đây accuracy tôi là bao nhiêu?", "C3: since_days=3 boundary", "aggregate_wrong_answers", False),

    # ── analyze_mistake (11) ─────────────────────────────────────────────────
    (2, "Câu hỏi: 'She ___ (go) to school'. Tôi điền 'goed'. Sai ở đâu?", "Analyze single mistake", "analyze_mistake", False),
    (2, "Phân tích lỗi giúp tôi: 'He have been to Paris', đáp án đúng 'has been'.", "Analyze subject-verb", "analyze_mistake", False),
    (3, "Câu 'I go to school yesterday' tôi viết vậy, sai chỗ nào?", "Analyze tense", "analyze_mistake", False),
    (3, "Đề: 'They ___ (finish) their homework'. Tôi trả lời 'finish'. Đúng là 'finished'. Giải thích.", "Past simple", "analyze_mistake", False),
    (4, "Tại sao 'She don't like coffee' là sai?", "Third person singular", "analyze_mistake", False),
    (4, "Câu 'If I was you, I would go' tôi viết vậy, chỗ nào chưa ổn?", "Conditional", "analyze_mistake", False),
    (2, "Câu hỏi 'The book ___ on the table' tôi viết 'are', đúng là 'is'. Vì sao?", "Subject-verb agreement", "analyze_mistake", False),
    (2, "Phân tích lỗi câu: 'He can plays football'. Tôi trả lời 'plays'.", "Modal verb", "analyze_mistake", False),
    # Nhóm B
    (3, "câu 'Me and him went there' này của t sai chỗ nào z", "Informal analyze", "analyze_mistake", False),
    # Nhóm C
    (4, "Đề bài 'She ___ TV every night'. T viết 'watch'. Đáp án 'watches'. Sai đâu?", "C3: full context param", "analyze_mistake", False),
    (2, "Tôi viết 'much people' — chỗ nào sai?", "C1: câu ngắn, thiếu context đề bài", "analyze_mistake", False),

    # ── get_recent_mistakes (10) ─────────────────────────────────────────────
    (2, "Tôi hay mắc lỗi ngữ pháp gì nhất?", "Recent grammar mistakes", "get_recent_mistakes", False),
    (2, "Lỗi gần đây của tôi là gì?", "Recent mistakes", "get_recent_mistakes", False),
    (3, "Điểm yếu ngữ pháp của tôi là gì?", "Weakness by grammar", "get_recent_mistakes", False),
    (3, "Nhắc lại những lỗi tôi đã phạm gần đây.", "Recall mistakes", "get_recent_mistakes", False),
    (4, "Tôi thường sai về ngữ pháp nào?", "Grammar pattern", "get_recent_mistakes", False),
    (4, "Cho tôi xem 10 lỗi gần nhất của tôi.", "Recent limit=10", "get_recent_mistakes", False),
    (2, "Tôi đã mắc những lỗi gì trong thời gian qua?", "Recent mistakes general", "get_recent_mistakes", False),
    # Nhóm B
    (3, "t hay bị lỗi vậy ta, xem hộ đi", "Slang", "get_recent_mistakes", False),
    # Nhóm C
    (4, "Không cần thống kê, chỉ liệt kê từng lỗi ngữ pháp tôi hay mắc.", "C2: phủ định + list mistakes", "get_recent_mistakes", False),
    (2, "phân tích lỗi này giúp tôi", "C1: ambiguous — không có câu cụ thể → fallback list", "get_recent_mistakes", False),

    # ── generate_passage (11) ────────────────────────────────────────────────
    (2, "Tạo cho tôi một đoạn văn luyện đọc về chủ đề travel.", "Passage travel", "generate_passage", False),
    (2, "Sinh bài đọc tiếng Anh về công nghệ.", "Passage tech", "generate_passage", False),
    (3, "Tôi muốn luyện đọc hiểu về môi trường.", "Passage environment", "generate_passage", False),
    (3, "Cho tôi bài đọc về ẩm thực kèm câu hỏi.", "Passage food + Q", "generate_passage", False),
    (4, "Tạo reading passage về business cho tôi.", "Passage business", "generate_passage", False),
    (4, "Sinh đoạn văn về sức khoẻ, có câu hỏi đọc hiểu.", "Passage health", "generate_passage", False),
    (2, "Tôi muốn 1 đoạn văn về giáo dục để luyện đọc.", "Passage education", "generate_passage", False),
    # Nhóm B
    (3, "cho t 1 bài đọc chủ đề du lịch đi", "Informal passage", "generate_passage", False),
    (4, "em ơi cho em xin bài đọc về công nghệ đi ạ", "Polite thân mật", "generate_passage", False),
    # Nhóm C
    (2, "Bài đọc về travel, 5 câu hỏi.", "C3: question_count=5 boundary", "generate_passage", False),
    (3, "cho bài đọc", "C1: ambiguous — không topic, expect default", "generate_passage", False),

    # ── recommend_questions (10) ─────────────────────────────────────────────
    (2, "Cho tôi vài câu hỏi phù hợp để luyện tập.", "Recommend basic", "recommend_questions", False),
    (2, "Gợi ý bài tập phù hợp với trình độ của tôi.", "Recommend by level", "recommend_questions", False),
    (3, "Tôi muốn luyện thêm câu hỏi về Grammar.", "Recommend by topic", "recommend_questions", False),
    (3, "Câu nào tôi nên luyện thêm?", "Recommend simple", "recommend_questions", False),
    (4, "Cho tôi 5 câu luyện tập topic Vocabulary.", "Recommend topic + limit", "recommend_questions", False),
    (4, "Recommend bài tập cho tôi.", "Recommend English word", "recommend_questions", False),
    (2, "Tôi muốn practice thêm, gợi ý câu hỏi giúp.", "Practice more", "recommend_questions", False),
    # Nhóm B
    (3, "cho vài câu practice đi bro", "Slang", "recommend_questions", False),
    # Nhóm C
    (4, "Cho 10 câu luyện tập topic 2.", "C3: topic_id=2 limit=10", "recommend_questions", False),
    (2, "Tôi rảnh, đưa vài bài ra đây.", "C1: ambiguous — 'đưa bài'", "recommend_questions", False),

    # ── explain_word_in_context (12) ─────────────────────────────────────────
    (2, "Trong câu 'The deal fell through at the last minute', 'fell through' nghĩa là gì?", "Explain phrasal verb", "explain_word_in_context", False),
    (2, "Từ 'bear' trong câu 'I can't bear this situation' có nghĩa gì?", "Contextual meaning", "explain_word_in_context", False),
    (3, "Giải thích 'run' trong câu 'She runs the whole department'.", "Verb in context", "explain_word_in_context", False),
    (3, "Từ 'draw' ở đây có nghĩa gì: 'The match ended in a draw'?", "Noun meaning", "explain_word_in_context", False),
    (4, "Trong câu 'He was fired from the company', 'fired' nghĩa là gì?", "Fired meaning", "explain_word_in_context", False),
    (4, "'Light' trong câu 'This box is very light' nghĩa là gì?", "Adjective in context", "explain_word_in_context", False),
    (2, "Trong 'She broke the news to me', 'broke' có nghĩa gì?", "Break contextual", "explain_word_in_context", False),
    (2, "Giải thích từ 'address' trong câu 'We need to address this issue'.", "Verb address", "explain_word_in_context", False),
    (3, "Trong câu 'He passed away last year', 'passed away' nghĩa là gì?", "Euphemism", "explain_word_in_context", False),
    # Nhóm B
    (4, "'get away with' trong 'He got away with cheating' là sao z", "Slang explain", "explain_word_in_context", False),
    (2, "em ơi trong câu 'time flies' thì 'flies' nghĩa gì ạ", "Polite thân mật", "explain_word_in_context", False),
    # Nhóm C
    (3, "Trong 'She takes after her mother', 'takes after' — 3 nghĩa gần nhất?", "C3: ép nhiều nghĩa", "explain_word_in_context", False),

    # ── get_lesson_detail (10) ───────────────────────────────────────────────
    (2, "Cho tôi xem chi tiết lesson 2.", "Detail lesson 2", "get_lesson_detail", False),
    (2, "Lesson 3 có những bài tập gì?", "Exercises in lesson 3", "get_lesson_detail", False),
    (3, "Bài học số 5 có nội dung gì?", "Lesson 5 content", "get_lesson_detail", False),
    (3, "Exercise trong lesson 4 là gì?", "Exercise list", "get_lesson_detail", False),
    (4, "Chi tiết bài 6 giúp tôi.", "Detail lesson 6", "get_lesson_detail", False),
    (4, "Lesson 1 khó không, có bao nhiêu exercise?", "Difficulty check lesson 1", "get_lesson_detail", False),
    (2, "Mô tả lesson 8 cho tôi.", "Describe lesson 8", "get_lesson_detail", False),
    # Nhóm B
    (3, "lesson 2 có gì chơi ta", "Slang", "get_lesson_detail", False),
    # Nhóm C
    (4, "Chi tiết lesson 7.", "C3: lesson_id=7 boundary", "get_lesson_detail", False),
    (2, "Cho biết lesson 10 có mấy exercise.", "C3: lesson_id=10", "get_lesson_detail", False),

    # ── search_snippet (10) ──────────────────────────────────────────────────
    (2, "Tìm clip audio về chủ đề job interview cho tôi.", "Snippet job interview", "search_snippet", False),
    (2, "Tôi muốn nghe ví dụ thực tế về cách dùng 'moreover'.", "Snippet moreover", "search_snippet", False),
    (3, "Có đoạn audio nào về daily conversation không?", "Snippet daily", "search_snippet", False),
    (3, "Tìm snippet về business English cho tôi.", "Snippet business", "search_snippet", False),
    (4, "Tìm ví dụ nghe về topic travel.", "Snippet travel", "search_snippet", False),
    (4, "Tìm snippet chứa từ 'negotiation'.", "Snippet negotiation", "search_snippet", False),
    (2, "Cho tôi audio clip về chủ đề ordering food.", "Snippet ordering food", "search_snippet", False),
    # Nhóm B
    (3, "kiếm clip nghe về interview đi", "Slang", "search_snippet", False),
    # Nhóm C
    (4, "Bỏ qua vocabulary lookup, tìm snippet nghe về 'moreover'.", "C2: phủ định + snippet", "search_snippet", False),
    (2, "Tìm 8 audio về chủ đề travel.", "C3: limit=8 boundary", "search_snippet", False),

    # ── MULTI-TOOL COMBOS (15) ───────────────────────────────────────────────
    # Cột "Tool expected" là tool CHÍNH agent phải gọi (tool tạo output cuối).
    (2, "Tôi đang ở level nào và nên làm dạng bài gì phù hợp?", "Progress + recommend", "recommend_questions", True),
    (2, "Kiểm tra tiến độ và gợi ý bài tập phù hợp cho tôi.", "Progress + recommend", "recommend_questions", True),
    (3, "Tôi hay sai gì gần đây, cho vài câu luyện tập lại đi.", "Recent mistakes + recommend", "recommend_questions", True),
    (3, "Xem accuracy tuần qua và đề xuất bài tập.", "Aggregate + recommend", "recommend_questions", True),
    (4, "Cho tôi biết level và tạo đoạn văn phù hợp.", "Progress + passage", "generate_passage", True),
    (4, "Tôi yếu topic nào? Sau đó gợi ý bài đọc phù hợp.", "Progress + passage", "generate_passage", True),
    (2, "Tra từ 'ambitious' và tìm audio clip có từ này.", "Lookup + snippet", "search_snippet", True),
    (2, "Từ 'commitment' nghĩa là gì và có clip nào chứa nó không?", "Lookup + snippet", "search_snippet", True),
    (3, "Xem chi tiết lesson 3 và cho tôi vài câu luyện tập cho lesson đó.", "Lesson detail + recommend", "recommend_questions", True),
    (3, "Liệt kê các topic có sẵn rồi cho tôi chi tiết lesson 1.", "Topics + lesson detail", "get_lesson_detail", True),
    (4, "Tôi hay sai câu nào? Kèm theo phân tích lỗi ngữ pháp gần nhất giúp tôi.", "Aggregate + recent mistakes", "get_recent_mistakes", True),
    (4, "Cho tôi biết trình độ hiện tại và lịch sử làm bài tuần qua.", "Progress + history", "get_user_answer_history", True),
    (2, "Xem accuracy của tôi và gợi ý topic tôi nên luyện.", "Aggregate + recommend topic", "recommend_questions", True),
    (3, "Tra từ 'persevere' rồi giải thích nghĩa của nó trong câu 'She persevered despite the setbacks'.", "Lookup + explain context", "explain_word_in_context", True),
    (4, "Cho tôi tiến độ hiện tại và một bài đọc phù hợp với trình độ đó.", "Progress + passage", "generate_passage", True),
]


def _sample_request(learner_id: int, message: str) -> str:
    return json.dumps(
        {
            "learner_id": learner_id,
            "messages": [{"role": "user", "content": message}],
        },
        ensure_ascii=False,
        indent=2,
    )


def build() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (lid, msg, purpose, tool, multi) in enumerate(CASES, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=ENDPOINT)
        ws.cell(row=row, column=3, value=lid)
        ws.cell(row=row, column=4, value=msg)
        ws.cell(row=row, column=5, value=purpose)
        ws.cell(row=row, column=6, value=tool)
        ws.cell(row=row, column=7, value="Yes" if multi else "No")
        ws.cell(row=row, column=8, value=_sample_request(lid, msg))
        for col in range(1, 9):
            ws.cell(row=row, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    widths = [6, 18, 12, 55, 40, 32, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print(f"Wrote {len(CASES)} cases to {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
