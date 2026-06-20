"""
Quality test for POST /grammar/chat — đánh giá chatbot theo 7 tiêu chí:
  1. Smoke   : HTTP 200, answer không rỗng
  2. Latency : thời gian phản hồi < LATENCY_THRESHOLD_S
  3. Length  : số từ trong khoảng [MIN_WORDS, MAX_WORDS]
  4. Accuracy: LLM-as-judge chấm độ chính xác nội dung (1-5)
  5. Relevance: LLM-as-judge chấm độ liên quan tới câu hỏi (1-5)
  6. Context : có nhận diện đúng exercise đang học không
  7. Consistency: chạy lặp N lần, đo độ ổn định (chỉ với --repeat)

Pytest mode (smoke + latency + length, không gọi judge):
    pytest tests/test_grammar_chat_smoke.py -v

Script mode (đầy đủ tiêu chí, ghi xlsx):
    python tests/test_grammar_chat_smoke.py
    python tests/test_grammar_chat_smoke.py --eval        # bật LLM judge
    python tests/test_grammar_chat_smoke.py --repeat 3    # đo consistency
    python tests/test_grammar_chat_smoke.py --multi-turn  # bật multi-turn case
"""

import argparse
import json
import statistics
import sys
import time
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
import pytest
from fastapi.testclient import TestClient
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.exc import SQLAlchemyError
from sqlmodel import Session, text

from app.database import engine
from app.main import app

XLSX_PATH = Path(__file__).parent / "chat_test_questions.xlsx"
SHEET_NAME = "grammar_chat"
ANSWER_PREVIEW_LEN = 120

# ── Thresholds (tinh chỉnh theo yêu cầu dự án) ───────────────────────────────
LATENCY_THRESHOLD_S = 15.0  # cảnh báo nếu vượt
LATENCY_HARD_S = 30.0  # FAIL nếu vượt
MIN_WORDS = 30  # quá ngắn → thiếu thông tin
MAX_WORDS = 800  # quá dài → mất tập trung
JUDGE_PASS_THRESHOLD = 3  # 1-5; <3 coi là FAIL về chất lượng

# ── Colors ──────────────────────────────────────────────────────────────────
_PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
_WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


# ── Data loading ─────────────────────────────────────────────────────────────


def _load_cases():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        row_num = int(row[0])
        learner_id = int(row[2])
        exercise_id = int(row[3])
        exercise_name = row[4] or ""
        message = row[5] or ""
        request_json_str = row[6] or ""
        try:
            request_body = json.loads(request_json_str)
        except (json.JSONDecodeError, TypeError):
            request_body = None
        cases.append(
            {
                "row_num": row_num,
                "learner_id": learner_id,
                "exercise_id": exercise_id,
                "exercise_name": exercise_name,
                "message": message,
                "request_body": request_body,
            }
        )
    wb.close()
    return cases


def load_grammar_chat_cases():
    return [
        pytest.param(
            c["learner_id"],
            c["message"],
            c["request_body"],
            c["exercise_name"],
            id=f"#{c['row_num']} {c['exercise_name'][:30]}",
        )
        for c in _load_cases()
    ]


# ── Core call (đo latency) ───────────────────────────────────────────────────


def _build_default_body(learner_id: int, message: str) -> dict:
    return {
        "learner_id": learner_id,
        "session_id": None,
        "messages": [{"role": "user", "content": message}],
        "context": {
            "exercise_id": 1,
            "exercise_name": "Grammar",
            "current_question_id": None,
            "questions": [],
        },
    }


def _call_chat(
    client: TestClient,
    learner_id: int,
    message: str,
    request_body: dict | None,
):
    """Gọi /grammar/chat. Trả về dict chứa kết quả + latency."""
    body = (
        request_body
        if request_body
        else _build_default_body(learner_id, message)
    )
    t0 = time.perf_counter()
    try:
        response = client.post("/grammar/chat", json=body)
        latency = time.perf_counter() - t0
        if response.status_code != 200:
            return {
                "ok": False,
                "answer": "",
                "latency": latency,
                "error": f"HTTP {response.status_code}: {response.text[:200]}",
            }
        data = response.json()
        answer = (data.get("answer") or "").strip()
        if not answer:
            return {
                "ok": False,
                "answer": "",
                "latency": latency,
                "error": "answer is empty",
            }
        return {"ok": True, "answer": answer, "latency": latency, "error": ""}
    except Exception as exc:
        return {
            "ok": False,
            "answer": "",
            "latency": time.perf_counter() - t0,
            "error": str(exc),
        }


# ── Quality metrics ──────────────────────────────────────────────────────────


def count_words(text_str: str) -> int:
    return len(text_str.split()) if text_str else 0


def length_status(words: int) -> str:
    """OK / TOO_SHORT / TOO_LONG."""
    if words < MIN_WORDS:
        return "TOO_SHORT"
    if words > MAX_WORDS:
        return "TOO_LONG"
    return "OK"


_STOPWORDS_VI = {
    "bài",
    "tập",
    "thì",
    "hiện",
    "tại",
    "quá",
    "khứ",
    "tương",
    "lai",
    "đơn",
    "tiếp",
    "diễn",
    "hoàn",
    "thành",
    "sự",
    "hòa",
    "hợp",
    "giữa",
    "các",
    "viết",
    "hình",
    "thức",
    "đúng",
    "của",
    "và",
    "để",
    "là",
}


def _exercise_keywords(exercise_name: str) -> list[str]:
    """Trích token tiếng Anh + tiếng Việt có ý nghĩa từ tên bài."""
    if not exercise_name:
        return []
    raw = exercise_name.lower().replace("–", " ").replace("-", " ")
    tokens = [t.strip(",.()") for t in raw.split() if t.strip()]
    return [t for t in tokens if len(t) >= 3 and t not in _STOPWORDS_VI]


def context_recognized(answer: str, exercise_name: str) -> bool:
    """Heuristic: chatbot có nhắc đến chủ đề bài hiện tại không?"""
    if not answer or not exercise_name:
        return False
    answer_lower = answer.lower()
    keywords = _exercise_keywords(exercise_name)
    if not keywords:
        return True
    hits = sum(1 for kw in keywords if kw in answer_lower)
    return hits >= max(1, len(keywords) // 3)


# ── LLM-as-judge ─────────────────────────────────────────────────────────────

_JUDGE_PROMPT = """Bạn là chuyên gia ngữ pháp tiếng Anh, đang chấm điểm phản hồi của một chatbot dạy ngữ pháp.

[Bài tập đang học]: {exercise_name}
[Câu hỏi của học viên]: {question}
[Phản hồi của chatbot]:
{answer}

Hãy chấm điểm theo 3 tiêu chí, mỗi tiêu chí từ 1 (rất kém) đến 5 (xuất sắc):
- accuracy: nội dung ngữ pháp/ví dụ có chính xác không
- relevance: phản hồi có trả lời đúng câu hỏi của học viên không
- on_topic: có nhận diện và phù hợp với bài tập đang học không (5 = nhận diện rõ;
  3 = trả lời đúng câu hỏi nhưng không nhắc bài hiện tại; 1 = lạc đề hoàn toàn)

Trả về DUY NHẤT một JSON object với 4 trường:
{{"accuracy": <int>, "relevance": <int>, "on_topic": <int>, "note": "<lý do ngắn gọn, <=200 ký tự>"}}
Không thêm bất kỳ markdown hay text nào khác.
"""


def llm_judge(question: str, answer: str, exercise_name: str) -> dict:
    """Trả về {accuracy, relevance, on_topic, note} hoặc dict lỗi."""
    try:
        from app.services.llm_service import call_llm
    except Exception as exc:
        return {
            "accuracy": 0,
            "relevance": 0,
            "on_topic": 0,
            "note": f"judge import error: {exc}",
        }

    prompt = _JUDGE_PROMPT.format(
        exercise_name=exercise_name or "Grammar",
        question=question,
        answer=answer,
    )
    try:
        raw = call_llm(prompt) or ""
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.strip("`")
            if cleaned.lower().startswith("json"):
                cleaned = cleaned[4:]
            cleaned = cleaned.strip()
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start == -1 or end == -1:
            return {
                "accuracy": 0,
                "relevance": 0,
                "on_topic": 0,
                "note": f"non-json: {raw[:120]}",
            }
        data = json.loads(cleaned[start : end + 1])
        return {
            "accuracy": int(data.get("accuracy", 0)),
            "relevance": int(data.get("relevance", 0)),
            "on_topic": int(data.get("on_topic", 0)),
            "note": str(data.get("note", ""))[:200],
        }
    except Exception as exc:
        return {
            "accuracy": 0,
            "relevance": 0,
            "on_topic": 0,
            "note": f"judge error: {exc}",
        }


# ── Consistency (chạy lặp) ───────────────────────────────────────────────────


def _jaccard(a: str, b: str) -> float:
    sa = set(a.lower().split())
    sb = set(b.lower().split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def consistency_score(answers: list[str]) -> float:
    """Trung bình Jaccard pairwise giữa các lần chạy. 1.0 = giống hệt."""
    if len(answers) < 2:
        return 1.0
    scores = []
    for i in range(len(answers)):
        for j in range(i + 1, len(answers)):
            scores.append(_jaccard(answers[i], answers[j]))
    return sum(scores) / len(scores) if scores else 0.0


# ── Multi-turn cases ─────────────────────────────────────────────────────────

MULTI_TURN_CASES = [
    {
        "name": "follow-up: tại sao vậy",
        "exercise_name": "Bài tập thì hiện tại đơn",
        "turns": [
            "Câu 2 tôi điền 'plays' nhưng sai.",
            "Tại sao vậy?",
        ],
    },
    {
        "name": "follow-up: ví dụ thêm",
        "exercise_name": "Bài tập thì hiện tại tiếp diễn",
        "turns": [
            "Quy tắc thêm -ing là gì?",
            "Cho tôi thêm vài ví dụ về động từ kết thúc bằng 'e'.",
        ],
    },
    {
        "name": "follow-up: làm rõ",
        "exercise_name": "Bài tập thì hiện tại hoàn thành",
        "turns": [
            "Khi nào dùng present perfect?",
            "Còn 'for' và 'since' khác nhau thế nào?",
        ],
    },
]


def _run_multi_turn(client: TestClient, case: dict) -> dict:
    """Gửi từng lượt, giữ nguyên session_id để test chatbot có nhớ context không."""
    session_id = None
    messages = []
    answers = []
    total_latency = 0.0
    error = ""
    for turn in case["turns"]:
        messages.append({"role": "user", "content": turn})
        body = {
            "learner_id": 1,
            "session_id": session_id,
            "messages": messages,
            "context": {
                "exercise_id": 1,
                "exercise_name": case["exercise_name"],
                "current_question_id": None,
                "questions": [],
            },
        }
        t0 = time.perf_counter()
        resp = client.post("/grammar/chat", json=body)
        total_latency += time.perf_counter() - t0
        if resp.status_code != 200:
            error = f"HTTP {resp.status_code}"
            break
        data = resp.json()
        ans = (data.get("answer") or "").strip()
        if not ans:
            error = f"empty answer at turn {len(answers) + 1}"
            break
        answers.append(ans)
        messages.append({"role": "assistant", "content": ans})
        session_id = data.get("session_id") or session_id
    return {
        "ok": not error and len(answers) == len(case["turns"]),
        "answers": answers,
        "latency": total_latency,
        "error": error,
    }


# ── Pytest mode (smoke + latency + length, không gọi judge) ─────────────────


@pytest.fixture(scope="module")
def grammar_client():
    try:
        with Session(engine) as s:
            s.exec(text("SELECT 1"))
    except SQLAlchemyError as exc:
        pytest.skip(f"DB unavailable: {exc}")
    with TestClient(app) as client:
        yield client


@pytest.mark.parametrize(
    "learner_id,message,request_body,exercise_name",
    load_grammar_chat_cases(),
)
def test_grammar_chat_quality(
    grammar_client, learner_id, message, request_body, exercise_name
):
    """Smoke + latency + length, không phụ thuộc LLM judge để pytest chạy nhanh."""
    result = _call_chat(grammar_client, learner_id, message, request_body)
    assert result["ok"], result["error"]
    assert result["latency"] < LATENCY_HARD_S, (
        f"Latency {result['latency']:.2f}s vượt ngưỡng cứng {LATENCY_HARD_S}s"
    )
    words = count_words(result["answer"])
    assert words >= MIN_WORDS, f"Phản hồi quá ngắn: {words} từ"
    assert words <= MAX_WORDS, f"Phản hồi quá dài: {words} từ"


@pytest.mark.parametrize(
    "case", MULTI_TURN_CASES, ids=[c["name"] for c in MULTI_TURN_CASES]
)
def test_grammar_chat_multi_turn(grammar_client, case):
    result = _run_multi_turn(grammar_client, case)
    assert result["ok"], result["error"]
    assert all(count_words(a) >= MIN_WORDS for a in result["answers"]), (
        "Có lượt trả lời quá ngắn trong multi-turn"
    )


# ── Script mode: xlsx layout ─────────────────────────────────────────────────

# Cột H trở đi (col 8+) là kết quả đánh giá
RESULT_COLUMNS = [
    ("Pass/Fail", 12),
    ("Latency (s)", 11),
    ("Words", 8),
    ("Length", 11),
    ("Context OK", 11),
    ("Judge Acc", 10),
    ("Judge Rel", 10),
    ("Judge Topic", 11),
    ("Consistency", 12),
    ("Answer preview", 50),
    ("Error", 35),
    ("Judge note", 40),
    ("Full answer", 80),
]
# Mapping tiện cho code dưới
COL_PASS = 8
COL_LATENCY = 9
COL_WORDS = 10
COL_LENGTH = 11
COL_CONTEXT = 12
COL_ACC = 13
COL_REL = 14
COL_TOPIC = 15
COL_CONSIST = 16
COL_PREVIEW = 17
COL_ERROR = 18
COL_NOTE = 19
COL_FULL = 20


def _init_result_columns(ws):
    for i, (name, width) in enumerate(RESULT_COLUMNS, start=COL_PASS):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width


def _row_fill(
    passed: bool, latency: float, words: int, judge: dict | None
) -> PatternFill:
    if not passed:
        return _FAIL_FILL
    if latency > LATENCY_THRESHOLD_S:
        return _WARN_FILL
    if words < MIN_WORDS or words > MAX_WORDS:
        return _WARN_FILL
    if judge:
        if (
            min(judge.get("accuracy", 5), judge.get("relevance", 5))
            < JUDGE_PASS_THRESHOLD
        ):
            return _WARN_FILL
    return _PASS_FILL


def _write_row(
    ws,
    row_num: int,
    *,
    passed: bool,
    latency: float,
    answer: str,
    error: str,
    exercise_name: str,
    judge: dict | None,
    consistency: float | None,
):
    xlsx_row = row_num + 1
    words = count_words(answer)
    fill = _row_fill(passed, latency, words, judge)
    preview = answer.strip()[:ANSWER_PREVIEW_LEN]

    values = {
        COL_PASS: "PASS" if passed else "FAIL",
        COL_LATENCY: round(latency, 3),
        COL_WORDS: words,
        COL_LENGTH: length_status(words),
        COL_CONTEXT: "YES"
        if context_recognized(answer, exercise_name)
        else "NO",
        COL_ACC: judge["accuracy"] if judge else "",
        COL_REL: judge["relevance"] if judge else "",
        COL_TOPIC: judge["on_topic"] if judge else "",
        COL_CONSIST: round(consistency, 3) if consistency is not None else "",
        COL_PREVIEW: preview,
        COL_ERROR: error,
        COL_NOTE: judge["note"] if judge else "",
        COL_FULL: answer,
    }
    for col, val in values.items():
        cell = ws.cell(row=xlsx_row, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    pf_cell = ws.cell(row=xlsx_row, column=COL_PASS)
    pf_cell.font = Font(bold=True, color="375623" if passed else "9C0006")


def _load_done_rows(out_path: Path) -> set[int]:
    if not out_path.exists():
        return set()
    wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        pf = row[COL_PASS - 1] if len(row) >= COL_PASS else None
        if pf in ("PASS", "FAIL"):
            done.add(int(row[0]))
    wb.close()
    return done


# ── Script mode: main loop ──────────────────────────────────────────────────


def _evaluate_case(
    client: TestClient, case: dict, *, eval_judge: bool, repeat: int
):
    """Chạy một case (có thể lặp), trả về kết quả tổng hợp."""
    runs = []
    for _ in range(max(1, repeat)):
        runs.append(
            _call_chat(
                client,
                case["learner_id"],
                case["message"],
                case["request_body"],
            )
        )

    primary = runs[0]
    answers_ok = [r["answer"] for r in runs if r["ok"]]
    consistency = (
        consistency_score(answers_ok) if len(answers_ok) >= 2 else None
    )

    judge = None
    if eval_judge and primary["ok"]:
        judge = llm_judge(
            case["message"], primary["answer"], case["exercise_name"]
        )

    return {
        "passed": primary["ok"],
        "latency": primary["latency"],
        "answer": primary["answer"],
        "error": primary["error"],
        "judge": judge,
        "consistency": consistency,
    }


def _print_case_log(case: dict, result: dict):
    status = "PASS" if result["passed"] else "FAIL"
    extra = f"{result['latency']:5.2f}s | {count_words(result['answer']):4d}w"
    if result["judge"]:
        j = result["judge"]
        extra += (
            f" | acc={j['accuracy']} rel={j['relevance']} top={j['on_topic']}"
        )
    if result["consistency"] is not None:
        extra += f" | consist={result['consistency']:.2f}"
    print(f"[{status}] #{case['row_num']:>3} {extra} | {case['message'][:50]}")
    if result["error"]:
        print(f"       ERROR: {result['error']}")


def _print_summary(results: list[dict]):
    total = len(results)
    if not total:
        print("No results.")
        return
    passed = sum(1 for r in results if r["passed"])
    latencies = [r["latency"] for r in results if r["passed"]]
    word_counts = [count_words(r["answer"]) for r in results if r["passed"]]

    print("\n" + "=" * 60)
    print(f"Tổng: {passed}/{total} PASS ({100 * passed / total:.1f}%)")
    if latencies:
        latencies_sorted = sorted(latencies)
        p50 = latencies_sorted[len(latencies_sorted) // 2]
        p95 = latencies_sorted[int(len(latencies_sorted) * 0.95)]
        print(
            f"Latency  : avg={statistics.mean(latencies):.2f}s  "
            f"p50={p50:.2f}s  p95={p95:.2f}s  max={max(latencies):.2f}s"
        )
    if word_counts:
        print(
            f"Words    : avg={statistics.mean(word_counts):.0f}  "
            f"min={min(word_counts)}  max={max(word_counts)}"
        )
        too_short = sum(1 for w in word_counts if w < MIN_WORDS)
        too_long = sum(1 for w in word_counts if w > MAX_WORDS)
        print(f"  too_short={too_short}  too_long={too_long}")

    judges = [r["judge"] for r in results if r["judge"]]
    if judges:
        print(
            f"Judge    : n={len(judges)}  "
            f"acc={statistics.mean(j['accuracy'] for j in judges):.2f}  "
            f"rel={statistics.mean(j['relevance'] for j in judges):.2f}  "
            f"topic={statistics.mean(j['on_topic'] for j in judges):.2f}"
        )

    consistencies = [
        r["consistency"] for r in results if r["consistency"] is not None
    ]
    if consistencies:
        print(
            f"Consistency (Jaccard avg): {statistics.mean(consistencies):.3f}"
        )
    print("=" * 60)


def _run_script(
    out_path: Path, *, eval_judge: bool, repeat: int, multi_turn: bool
):
    try:
        with Session(engine) as s:
            s.exec(text("SELECT 1"))
    except SQLAlchemyError as exc:
        print(f"DB unavailable, aborting: {exc}")
        sys.exit(1)

    import shutil

    if not out_path.exists():
        shutil.copy(XLSX_PATH, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb[SHEET_NAME]
    _init_result_columns(ws)
    wb.save(out_path)

    done_rows = _load_done_rows(out_path)
    all_cases = _load_cases()
    pending = [c for c in all_cases if c["row_num"] not in done_rows]
    if done_rows:
        print(f"Resuming: {len(done_rows)} done, {len(pending)} pending")

    print(
        f"Config: eval_judge={eval_judge}  repeat={repeat}  multi_turn={multi_turn}"
    )
    results = []

    with TestClient(app) as client:
        for c in pending:
            res = _evaluate_case(
                client, c, eval_judge=eval_judge, repeat=repeat
            )
            _print_case_log(c, res)
            results.append(res)

            wb = openpyxl.load_workbook(out_path)
            ws = wb[SHEET_NAME]
            _write_row(
                ws,
                c["row_num"],
                passed=res["passed"],
                latency=res["latency"],
                answer=res["answer"],
                error=res["error"],
                exercise_name=c["exercise_name"],
                judge=res["judge"],
                consistency=res["consistency"],
            )
            wb.save(out_path)

        if multi_turn:
            print("\n--- Multi-turn cases ---")
            for case in MULTI_TURN_CASES:
                mt = _run_multi_turn(client, case)
                status = "PASS" if mt["ok"] else "FAIL"
                print(
                    f"[{status}] {case['name']} | turns={len(mt['answers'])}/"
                    f"{len(case['turns'])} | {mt['latency']:.2f}s | {mt['error']}"
                )

    _print_summary(results)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).parent
        / f"grammar_quality_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    parser.add_argument(
        "--eval", action="store_true", help="Bật LLM-as-judge (chậm hơn)"
    )
    parser.add_argument(
        "--repeat",
        type=int,
        default=1,
        help="Số lần gọi mỗi case để đo consistency",
    )
    parser.add_argument(
        "--multi-turn",
        action="store_true",
        help="Chạy thêm bộ multi-turn cases",
    )
    args = parser.parse_args()
    _run_script(
        args.output,
        eval_judge=args.eval,
        repeat=args.repeat,
        multi_turn=args.multi_turn,
    )
