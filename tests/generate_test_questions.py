"""
Generates test questions for:
  - POST /grammar/chat  (grammar chatbot)
  - POST /agent/chat    (general learning agent with tool calls)
and saves them to tests/chat_test_questions.xlsx

agent_chat sheet has 100 rows covering all 14 tools in tool_registry.py:
  get_user_progress, get_user_answer_history, get_topics_lesson,
  lookup_vocabulary, aggregate_wrong_answers, batch_analyze_wrong_answers,
  analyze_mistake, get_recent_mistakes, generate_passage, generate_study_plan,
  recommend_questions, explain_word_in_context, get_lesson_detail,
  search_snippet
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# (imports used below in write_excel and make_*_payload)

# ---------------------------------------------------------------------------
# 50 test cases for POST /grammar/chat
# ---------------------------------------------------------------------------
GRAMMAR_QUESTIONS = [
    # --- Basic comprehension ---
    (1, "Present Simple", 1, "Câu hỏi này hỏi về điều gì?"),
    (1, "Present Simple", 1, "Đáp án đúng của câu này là gì? Giải thích cho tôi."),
    (1, "Present Simple", 2, "Tại sao đáp án không phải là 'goes'?"),
    (1, "Present Simple", 2, "Cho tôi ví dụ thêm về thì hiện tại đơn."),
    (1, "Present Simple", 3, "Phân biệt 'do' và 'does' trong câu hỏi."),
    # --- Wrong answer follow-up ---
    (2, "Past Simple", 1, "Tôi trả lời sai câu này, tại sao vậy?"),
    (2, "Past Simple", 1, "Hướng dẫn tôi cách dùng 'was' và 'were'."),
    (2, "Past Simple", 2, "Ở câu số 3, tôi chọn 'goed' — lỗi này phổ biến không?"),
    (2, "Past Simple", 2, "Làm sao nhớ các động từ bất quy tắc?"),
    (2, "Past Simple", 3, "Câu hỏi có 'yesterday' thì dùng thì gì?"),
    # --- Hint / suggestion requests ---
    (3, "Present Continuous", 1, "Gợi ý cho tôi từ khóa để làm bài này."),
    (3, "Present Continuous", 1, "Câu này nên điền 'is playing' hay 'plays'?"),
    (3, "Present Continuous", 2, "Khi nào dùng '-ing' form?"),
    (3, "Present Continuous", 2, "Phân biệt present simple và present continuous."),
    (3, "Present Continuous", 3, "Cho tôi thêm bài tập tương tự để luyện."),
    # --- Explanation of grammar point ---
    (4, "Future Simple", 1, "Will và be going to khác nhau thế nào?"),
    (4, "Future Simple", 1, "Câu này dùng 'will' hay 'shall'?"),
    (4, "Future Simple", 2, "Viết lại câu này theo thì tương lai đơn."),
    (4, "Future Simple", 2, "Cho ví dụ về kế hoạch tương lai."),
    (4, "Future Simple", 3, "Tại sao không dùng present simple cho câu này?"),
    # --- Multiple choice reasoning ---
    (5, "Modal Verbs", 1, "Tại sao đáp án là 'should' chứ không phải 'must'?"),
    (5, "Modal Verbs", 1, "Can, could, may, might — dùng khi nào?"),
    (5, "Modal Verbs", 2, "Câu này test ability hay permission?"),
    (5, "Modal Verbs", 2, "Giải thích nghĩa của 'ought to'."),
    (5, "Modal Verbs", 3, "Modal verb nào dùng để đưa ra lời khuyên?"),
    # --- Fill-in-the-blank guidance ---
    (6, "Prepositions", 1, "Ô trống này điền giới từ gì?"),
    (6, "Prepositions", 1, "Phân biệt 'in', 'on', 'at' trong chỉ thời gian."),
    (6, "Prepositions", 2, "Sau 'interested' dùng giới từ gì?"),
    (6, "Prepositions", 2, "Cho ví dụ câu dùng 'by the end of'."),
    (6, "Prepositions", 3, "Tại sao không điền 'in' cho ô này?"),
    # --- Score / feedback questions ---
    (7, "Articles", 1, "Tôi làm đúng bao nhiêu câu rồi?"),
    (7, "Articles", 1, "Điểm yếu của tôi ở bài này là gì?"),
    (7, "Articles", 2, "Câu nào tôi mắc lỗi nhiều nhất?"),
    (7, "Articles", 2, "a, an, the — khi nào không dùng article?"),
    (7, "Articles", 3, "Giải thích lỗi sai ở câu số 2 của tôi."),
    # --- Contextual help ---
    (8, "Conditionals", 1, "Câu điều kiện loại 2 khác loại 1 thế nào?"),
    (8, "Conditionals", 1, "Viết lại câu này dưới dạng conditional type 3."),
    (8, "Conditionals", 2, "Tại sao dùng 'would have' ở đây?"),
    (8, "Conditionals", 2, "Mixed conditional là gì? Cho ví dụ."),
    (8, "Conditionals", 3, "If I were you — cấu trúc này thuộc loại mấy?"),
    # --- Vocabulary / word choice ---
    (9, "Reported Speech", 1, "Backshift tense trong reported speech là gì?"),
    (9, "Reported Speech", 1, "Câu trực tiếp này chuyển sang gián tiếp thế nào?"),
    (9, "Reported Speech", 2, "Tại sao 'today' đổi thành 'that day'?"),
    (9, "Reported Speech", 2, "Reporting verbs nào khác ngoài 'said'?"),
    (9, "Reported Speech", 3, "Câu hỏi Yes/No chuyển sang reported speech ra sao?"),
    # --- Review / summary ---
    (10, "Passive Voice", 1, "Câu bị động được hình thành như thế nào?"),
    (10, "Passive Voice", 1, "Tại sao câu này dùng bị động mà không dùng chủ động?"),
    (10, "Passive Voice", 2, "By + agent — khi nào cần, khi nào bỏ?"),
    (10, "Passive Voice", 2, "Đổi câu chủ động này sang bị động giúp tôi."),
    (10, "Passive Voice", 3, "Passive voice ở thì perfect là gì?"),
]

# ---------------------------------------------------------------------------
# 100 test cases for POST /agent/chat
# Each entry: (learner_id, user_message, description_of_intent, tool_expected, multi_tool)
# tool_expected: primary tool agent should call (or "no_tool" for pure knowledge)
# multi_tool: True if the question likely requires chaining 2+ tools
# ---------------------------------------------------------------------------
# Tools covered:
#   get_user_progress, get_user_answer_history, get_topics_lesson,
#   lookup_vocabulary, aggregate_wrong_answers, batch_analyze_wrong_answers,
#   analyze_mistake, get_recent_mistakes, generate_passage, generate_study_plan,
#   recommend_questions, explain_word_in_context, get_lesson_detail,
#   search_snippet
# ---------------------------------------------------------------------------

AGENT_QUESTIONS = [
    # ── get_user_progress (7 câu) ────────────────────────────────────────
    (1, "Trình độ hiện tại của tôi là gì?", "Hỏi level CEFR / theta",
     "get_user_progress", False),
    (1, "Theta của tôi đang ở mức bao nhiêu?", "Hỏi theta cụ thể",
     "get_user_progress", False),
    (2, "Tôi đã học được bao nhiêu lesson rồi?", "Đếm lesson đã học",
     "get_user_progress", False),
    (2, "Lesson nào tôi học tốt nhất, lesson nào yếu nhất?", "Phân tích theta từng lesson",
     "get_user_progress", False),
    (3, "Tôi đang ở level A2, B1 hay cao hơn?", "Kiểm tra CEFR label",
     "get_user_progress", False),
    (3, "Tiến độ tổng quan của tôi ra sao?", "Xem overview tiến độ",
     "get_user_progress", False),
    (1, "Tôi cần cải thiện điểm yếu nào để lên level tiếp theo?",
     "Phân tích điểm yếu + gợi ý", "get_user_progress", False),

    # ── get_user_answer_history (7 câu) ──────────────────────────────────
    (1, "Tôi đã làm bao nhiêu câu hỏi trong 7 ngày qua?",
     "Lịch sử 7 ngày gần đây", "get_user_answer_history", False),
    (1, "Accuracy của tôi tháng này là bao nhiêu?",
     "Accuracy tổng trong 30 ngày", "get_user_answer_history", False),
    (2, "Cho tôi xem 10 câu hỏi gần đây nhất tôi đã trả lời.",
     "Lịch sử câu hỏi limit=10", "get_user_answer_history", False),
    (2, "Kết quả làm bài ở lesson 5 của tôi thế nào?",
     "Lọc history theo lesson_id", "get_user_answer_history", False),
    (3, "Tuần qua tôi làm đúng bao nhiêu phần trăm?",
     "Accuracy since_days=7", "get_user_answer_history", False),
    (3, "Trong topic Grammar tôi có kết quả ra sao?",
     "Lọc history theo topic_id", "get_user_answer_history", False),
    (1, "Tôi đã trả lời đúng câu nào và sai câu nào trong bài kiểm tra vừa rồi?",
     "Chi tiết đúng/sai gần đây", "get_user_answer_history", False),

    # ── get_topics_lesson (5 câu) ─────────────────────────────────────────
    (1, "Hệ thống có những chủ đề nào để học?",
     "Liệt kê toàn bộ topic", "get_topics_lesson", False),
    (2, "Cho tôi xem danh sách các lesson hiện có.",
     "Liệt kê tất cả lesson", "get_topics_lesson", False),
    (2, "Có bao nhiêu topic ngữ pháp trong hệ thống?",
     "Đếm topic grammar", "get_topics_lesson", False),
    (3, "Tôi chưa học topic nào? Cho tôi xem các topic còn lại.",
     "So sánh topic hệ thống vs đã học", "get_topics_lesson", True),
    (3, "Liệt kê các bài học theo từng topic cho tôi.",
     "Cây topic → lesson", "get_topics_lesson", False),

    # ── lookup_vocabulary (7 câu) ─────────────────────────────────────────
    (1, "Tra từ 'resilience' cho tôi.",
     "Tra từ vựng cơ bản", "lookup_vocabulary", False),
    (1, "Cho tôi 3 ví dụ câu dùng từ 'ambiguous'.",
     "Lookup vocab với limit=3", "lookup_vocabulary", False),
    (2, "Từ 'mitigate' có những brick/câu ví dụ nào trong hệ thống?",
     "Tìm brick chứa từ", "lookup_vocabulary", False),
    (2, "Nghĩa và cách dùng của 'leverage' trong tiếng Anh kinh doanh?",
     "Lookup từ chuyên ngành", "lookup_vocabulary", False),
    (3, "Tìm ví dụ câu dùng cụm 'in spite of'.",
     "Lookup cụm từ", "lookup_vocabulary", False),
    (3, "Từ 'eloquent' xuất hiện trong những ngữ cảnh nào?",
     "Ngữ cảnh dùng từ", "lookup_vocabulary", False),
    (1, "Cho tôi tối đa 10 ví dụ về từ 'get' trong hệ thống.",
     "Lookup với limit=10", "lookup_vocabulary", False),

    # ── explain_word_in_context (6 câu) ──────────────────────────────────
    (1, "Trong câu 'The deal fell through at the last minute', 'fell through' nghĩa là gì?",
     "Giải thích phrasal verb trong ngữ cảnh", "explain_word_in_context", False),
    (1, "Câu 'She has a keen eye for detail' — 'keen' ở đây nghĩa là gì?",
     "Giải thích adjective trong ngữ cảnh", "explain_word_in_context", False),
    (2, "Trong đoạn văn 'The economy is picking up steam', 'picking up steam' nghĩa gì?",
     "Giải thích idiom trong ngữ cảnh", "explain_word_in_context", False),
    (2, "Từ 'will' trong câu 'He will have finished by noon' dùng với nghĩa gì?",
     "Giải thích modal trong future perfect", "explain_word_in_context", False),
    (3, "Trong câu 'I can't put my finger on it', 'put my finger on' nghĩa là gì?",
     "Giải thích idiom", "explain_word_in_context", False),
    (3, "'Run' trong câu 'Let's run through the agenda' nghĩa là gì?",
     "Giải thích phrasal verb", "explain_word_in_context", False),

    # ── aggregate_wrong_answers (6 câu) ──────────────────────────────────
    (1, "Tôi hay sai câu nào nhất? Cho tôi thống kê nhanh.",
     "Thống kê câu sai (không LLM)", "aggregate_wrong_answers", False),
    (1, "Accuracy của tôi ở lesson 3 là bao nhiêu?",
     "Accuracy theo lesson_id", "aggregate_wrong_answers", False),
    (2, "Tuần qua tôi sai bao nhiêu câu?",
     "Thống kê sai since_days=7", "aggregate_wrong_answers", False),
    (2, "Phân bố câu sai theo độ khó của tôi ra sao?",
     "Phân bố theo difficulty", "aggregate_wrong_answers", False),
    (3, "Tôi sai nhiều nhất ở topic nào trong 30 ngày qua?",
     "Sai theo topic + since_days=30", "aggregate_wrong_answers", False),
    (3, "Có bao nhiêu câu distinct mà tôi đã sai ít nhất 2 lần?",
     "Distinct wrong questions", "aggregate_wrong_answers", False),

    # ── batch_analyze_wrong_answers (5 câu) ──────────────────────────────
    (1, "Phân tích toàn bộ lỗi sai của tôi và cho tôi biết pattern lỗi.",
     "Batch phân tích lỗi → lưu memory", "batch_analyze_wrong_answers", False),
    (1, "Tôi muốn hiểu rõ tại sao mình hay sai. Phân tích giúp tôi.",
     "Request phân tích lỗi sâu", "batch_analyze_wrong_answers", False),
    (2, "Phân tích lỗi sai của tôi trong lesson 4 đi.",
     "Batch analyze lọc theo lesson", "batch_analyze_wrong_answers", False),
    (2, "Tôi vừa làm xong bài kiểm tra và sai nhiều câu. Phân tích lỗi cho tôi.",
     "Phân tích sau bài kiểm tra", "batch_analyze_wrong_answers", True),
    (3, "Phân tích top 5 câu sai nhiều nhất của tôi tháng này.",
     "Batch analyze limit=5 + since_days=30", "batch_analyze_wrong_answers", False),

    # ── analyze_mistake (6 câu) ───────────────────────────────────────────
    (1, "Câu hỏi: 'She ___ (go) to school'. Tôi điền 'goed'. Sai ở đâu vậy?",
     "Phân tích lỗi sai đơn lẻ", "analyze_mistake", False),
    (1, "Đề: 'He has been working here since 2010.' Tôi chọn 'worked' thay vì 'been working'. Giải thích lỗi.",
     "Phân tích lỗi present perfect continuous", "analyze_mistake", False),
    (2, "Câu: 'If I was you, I would do it.' Tôi viết 'was'. Đây có phải lỗi không?",
     "Phân tích lỗi conditional type 2", "analyze_mistake", False),
    (2, "Tôi viết 'The informations are useful.' Lỗi này là gì?",
     "Phân tích lỗi uncountable noun", "analyze_mistake", False),
    (3, "Câu đúng là 'Despite being tired, he finished'. Tôi viết 'Despite he was tired'. Sai chỗ nào?",
     "Phân tích lỗi despite + gerund", "analyze_mistake", False),
    (3, "Đề: chọn giới từ đúng sau 'interested'. Tôi chọn 'in'. Đúng hay sai? Giải thích.",
     "Phân tích lỗi preposition", "analyze_mistake", False),

    # ── get_recent_mistakes (5 câu) ───────────────────────────────────────
    (1, "Tôi hay sai gì gần đây nhất?",
     "Lấy lỗi sai từ MistakeMemory", "get_recent_mistakes", False),
    (1, "Cho tôi xem 10 lỗi sai gần nhất của tôi.",
     "MistakeMemory limit=10", "get_recent_mistakes", False),
    (2, "Thống kê lỗi của tôi theo loại lỗi (grammar point).",
     "Thống kê theo loại lỗi", "get_recent_mistakes", False),
    (2, "Tôi hay mắc loại lỗi ngữ pháp nào nhất?",
     "Pattern lỗi phổ biến", "get_recent_mistakes", False),
    (3, "Tóm tắt điểm yếu ngữ pháp của tôi dựa trên lịch sử lỗi.",
     "Tổng hợp điểm yếu từ MistakeMemory", "get_recent_mistakes", False),

    # ── generate_passage (6 câu) ──────────────────────────────────────────
    (1, "Tạo cho tôi một đoạn văn luyện đọc về chủ đề travel.",
     "Sinh passage topic travel", "generate_passage", False),
    (1, "Cho tôi đoạn văn về technology kèm 5 câu hỏi đọc hiểu.",
     "Passage + question_count=5", "generate_passage", False),
    (2, "Sinh đoạn văn về environment phù hợp với trình độ của tôi.",
     "Passage tự động điều chỉnh theo theta", "generate_passage", False),
    (2, "Tôi muốn luyện đọc về chủ đề business. Tạo bài cho tôi.",
     "Passage topic business", "generate_passage", False),
    (3, "Đoạn văn luyện đọc về health and medicine với 3 câu hỏi.",
     "Passage topic health + 3 questions", "generate_passage", False),
    (3, "Làm bài đọc hiểu về chủ đề science cho tôi.",
     "Passage topic science", "generate_passage", False),

    # ── generate_study_plan (6 câu) ───────────────────────────────────────
    (1, "Lập kế hoạch học TOEIC 600 trong 2 tháng cho tôi.",
     "Study plan TOEIC 600 / 8 tuần", "generate_study_plan", True),
    (1, "Tôi muốn đạt IELTS 6.0 trong 4 tháng. Lên kế hoạch giúp tôi.",
     "Study plan IELTS 6.0 / 16 tuần", "generate_study_plan", True),
    (2, "Lập kế hoạch học từ vựng trong 4 tuần.",
     "Study plan vocabulary / 4 tuần", "generate_study_plan", False),
    (2, "Tôi chỉ có 1 tháng để cải thiện grammar. Kế hoạch ra sao?",
     "Study plan grammar / 4 tuần", "generate_study_plan", True),
    (3, "Lên lịch học conversation English cho người mới bắt đầu trong 6 tuần.",
     "Study plan conversation / 6 tuần", "generate_study_plan", False),
    (3, "Kế hoạch ôn tập toàn diện trước kỳ thi TOEFL trong 3 tuần.",
     "Study plan TOEFL / 3 tuần", "generate_study_plan", True),

    # ── recommend_questions (6 câu) ───────────────────────────────────────
    (1, "Cho tôi vài câu hỏi phù hợp để luyện tập.",
     "Recommend 5 câu mặc định", "recommend_questions", False),
    (1, "Gợi ý 10 câu luyện tập cho tôi.",
     "Recommend limit=10", "recommend_questions", False),
    (2, "Cho tôi câu hỏi luyện tập trong topic Grammar.",
     "Recommend lọc theo topic_id", "recommend_questions", False),
    (2, "Tôi muốn luyện thêm câu ở mức độ vừa với trình độ hiện tại.",
     "Recommend theo theta hiện tại", "recommend_questions", False),
    (3, "Gợi ý 3 câu khó hơn mức hiện tại một chút để thách thức.",
     "Recommend limit=3", "recommend_questions", False),
    (3, "Cho tôi bài tập về topic Vocabulary để luyện.",
     "Recommend by topic vocabulary", "recommend_questions", False),

    # ── get_lesson_detail (5 câu) ─────────────────────────────────────────
    (1, "Cho tôi xem chi tiết lesson 2.",
     "Chi tiết lesson_id=2", "get_lesson_detail", False),
    (1, "Lesson 5 có những concept và exercise nào?",
     "Concept + exercise của lesson_id=5", "get_lesson_detail", False),
    (2, "Mô tả lesson 3 cho tôi nghe.",
     "Mô tả lesson_id=3", "get_lesson_detail", False),
    (2, "Lesson 7 gồm những phần nào? Tôi chưa bắt đầu.",
     "Chi tiết lesson trước khi học", "get_lesson_detail", False),
    (3, "Lesson 1 có bao nhiêu exercise? Liệt kê giúm tôi.",
     "Đếm exercise trong lesson_id=1", "get_lesson_detail", False),

    # ── search_snippet (5 câu) ────────────────────────────────────────────
    (1, "Tìm clip audio về chủ đề job interview cho tôi.",
     "Search snippet topic job interview", "search_snippet", False),
    (1, "Tìm snippet nào có từ 'environment' trong transcript.",
     "Search snippet keyword", "search_snippet", False),
    (2, "Cho tôi nghe ví dụ thực tế về cách dùng phrasal verb 'give up'.",
     "Search snippet phrasal verb", "search_snippet", False),
    (2, "Tìm clip audio về business meeting để tôi luyện nghe.",
     "Search snippet business meeting", "search_snippet", False),
    (3, "Có snippet nào về chủ đề daily conversation không?",
     "Search snippet daily conversation", "search_snippet", False),

    # ── Multi-tool scenarios (chaining 2+ tools) (10 câu) ─────────────────
    (1, "Tôi đang ở level nào và nên làm dạng bài gì phù hợp?",
     "get_user_progress → recommend_questions", "get_user_progress", True),
    (1, "Phân tích lỗi sai tuần qua rồi gợi ý bài tập phù hợp cho tôi.",
     "aggregate_wrong_answers → recommend_questions", "aggregate_wrong_answers", True),
    (2, "Dựa trên điểm yếu của tôi, lập kế hoạch học 4 tuần.",
     "get_recent_mistakes → generate_study_plan", "get_recent_mistakes", True),
    (2, "Tôi hay sai gì? Hãy phân tích và đề xuất đoạn văn luyện tập phù hợp.",
     "get_recent_mistakes → generate_passage", "get_recent_mistakes", True),
    (3, "Xem tiến độ của tôi rồi gợi ý topic tôi chưa học.",
     "get_user_progress → get_topics_lesson", "get_user_progress", True),
    (1, "Lấy lịch sử bài làm trong lesson 2 rồi cho tôi thấy chi tiết lesson đó.",
     "get_user_answer_history → get_lesson_detail", "get_user_answer_history", True),
    (2, "Tôi vừa sai câu 'He don't like it'. Phân tích lỗi rồi gợi ý bài luyện.",
     "analyze_mistake → recommend_questions", "analyze_mistake", True),
    (3, "Xem tôi học được gì rồi tìm snippet luyện nghe phù hợp.",
     "get_user_progress → search_snippet", "get_user_progress", True),
]


# ---------------------------------------------------------------------------
# Build ChatRequest JSON for grammar/chat
# ---------------------------------------------------------------------------
def make_grammar_payload(exercise_id, exercise_name, learner_id, question):
    return json.dumps(
        {
            "learner_id": learner_id,
            "session_id": None,
            "messages": [{"role": "user", "content": question}],
            "context": {
                "exercise_id": exercise_id,
                "exercise_name": exercise_name,
                "current_question_id": None,
                "questions": [
                    {
                        "order_id": 1,
                        "question_id": exercise_id * 10 + 1,
                        "question": "She ___ (go) to school every day.",
                        "user_answer": None,
                    }
                ],
            },
        },
        ensure_ascii=False,
        indent=2,
    )


# ---------------------------------------------------------------------------
# Build AgentChatRequest JSON for agent/chat
# ---------------------------------------------------------------------------
def make_agent_payload(learner_id, message):
    return json.dumps(
        {
            "learner_id": learner_id,
            "messages": [{"role": "user", "content": message}],
        },
        ensure_ascii=False,
        indent=2,
    )


# ---------------------------------------------------------------------------
# Write to Excel
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
GRAMMAR_FILL = PatternFill("solid", fgColor="D9E1F2")
AGENT_FILL = PatternFill("solid", fgColor="E2EFDA")
MULTI_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True)


def write_excel(out_path: Path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Grammar Chat ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "grammar_chat"
    headers1 = [
        "#", "Endpoint", "learner_id", "exercise_id", "exercise_name",
        "Câu hỏi người dùng", "Request JSON mẫu",
    ]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (ex_id, ex_name, l_id, q) in enumerate(GRAMMAR_QUESTIONS, 1):
        payload = make_grammar_payload(ex_id, ex_name, l_id, q)
        row = [i, "POST /grammar/chat", l_id, ex_id, ex_name, q, payload]
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=i + 1, column=col, value=val)
            cell.fill = GRAMMAR_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 22
    ws1.column_dimensions["F"].width = 45
    ws1.column_dimensions["G"].width = 55
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Agent Chat ────────────────────────────────────────────────
    ws2 = wb.create_sheet("agent_chat")
    headers2 = [
        "#", "Endpoint", "learner_id", "Câu hỏi người dùng",
        "Mục đích test", "Tool expected", "Multi-tool?",
        "Request JSON mẫu",
    ]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (l_id, msg, intent, tool, multi) in enumerate(AGENT_QUESTIONS, 1):
        payload = make_agent_payload(l_id, msg)
        row = [i, "POST /agent/chat", l_id, msg, intent, tool,
               "Yes" if multi else "No", payload]
        fill = MULTI_FILL if multi else AGENT_FILL
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=i + 1, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 52
    ws2.column_dimensions["E"].width = 38
    ws2.column_dimensions["F"].width = 30
    ws2.column_dimensions["G"].width = 12
    ws2.column_dimensions["H"].width = 55
    ws2.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  grammar/chat rows : {len(GRAMMAR_QUESTIONS)}")
    print(f"  agent/chat rows   : {len(AGENT_QUESTIONS)}")
    print(f"  Total             : {len(GRAMMAR_QUESTIONS) + len(AGENT_QUESTIONS)}")
    multi_count = sum(1 for *_, m in AGENT_QUESTIONS if m)
    print(f"  Multi-tool cases  : {multi_count}")


if __name__ == "__main__":
    out = Path(__file__).parent / "chat_test_questions.xlsx"
    write_excel(out)
